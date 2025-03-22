# text_extraction.py

import pdfplumber
import docx
import openpyxl
import pytesseract
import extract_msg
import email
from PIL import Image

def extract_text_from_pdf(pdf_path):
    """Extract text from a PDF file using pdfplumber."""
    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    return text

def extract_text_from_docx(docx_path):
    """Extract text from a DOCX file using python-docx."""
    doc = docx.Document(docx_path)
    return "\n".join(para.text for para in doc.paragraphs)

def extract_text_from_excel(excel_path):
    """Extract text from an Excel file using openpyxl."""
    wb = openpyxl.load_workbook(excel_path)
    text = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            text.append(" ".join(str(cell) for cell in row if cell))
    return "\n".join(text)

def extract_text_from_image(image_path):
    """
    Attempt to open and OCR an image. 
    If it's an unsupported format or Tesseract fails, skip it gracefully.
    """
    try:
        image = Image.open(image_path)
    except Exception as e:
        print(f"Skipping file '{image_path}' (not a valid image): {e}")
        return ""  # Return empty text if PIL can't open it

    try:
        return pytesseract.image_to_string(image)
    except Exception as e:
        print(f"Skipping file '{image_path}' (Tesseract error): {e}")
        return ""

def extract_text_from_email(msg_path):
    """Extract text from an Outlook MSG file using extract_msg."""
    msg = extract_msg.Message(msg_path)
    return (
        f"Subject: {msg.subject}\n"
        f"From: {msg.sender}\n"
        f"To: {msg.to}\n"
        f"Date: {msg.date}\n\n{msg.body}"
    )

def extract_text_from_eml(eml_path):
    """
    Extract basic text from a .eml file using Python's 'email' library.
    Returns a text representation including Subject, From, To, and the message body.
    """
    with open(eml_path, "r", encoding="utf-8", errors="ignore") as f:
        raw_data = f.read()

    # Parse the raw email data
    msg = email.message_from_string(raw_data)

    # Basic fields
    subject = msg.get("Subject", "")
    sender = msg.get("From", "")
    recipient = msg.get("To", "")
    
    text = f"Subject: {subject}\nFrom: {sender}\nTo: {recipient}\n\n"

    # If it's multipart, look for text/plain part(s)
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                payload = part.get_payload(decode=True)
                if payload:
                    text += payload.decode(errors="ignore")
    else:
        # Single-part email
        payload = msg.get_payload(decode=True)
        if payload:
            text += payload.decode(errors="ignore")

    return text