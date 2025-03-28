# excel_output.py

import os
import re
import openpyxl
import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

def remove_invalid_xml_chars(s):
    """
    Removes characters that are invalid in XML (ASCII control chars),
    allowing tab (\x09), newline (\x0A), and carriage return (\x0D).
    """
    if not isinstance(s, str):
        return s
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]", "", s)

def save_to_excel(df, output_filename):
    """
    Writes the DataFrame to Excel with:
      - A black header row at row 6
      - Data from row 7 onward
      - Adds specified text into specified cells

    'output_filename' is a full path to the final .xlsx file.
    """

    if df.empty:
        print("⚠️ No data to save. The DataFrame is empty.")
        return

    # Example custom headers, must match df.columns
    custom_headers = [
        "Document ID",
        "Document Date",
        "Category",
        "Document Title",
        "Summary",
        "Document Type",
        "Host Document ID",
        "Author",
        "Recipient",
        "Filename (including extension)",
        "Hyperlink"
    ]

    # Write the DataFrame at row=7 (header row=6 is black)
    with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
        df.to_excel(
            writer,
            sheet_name="Document Register",
            index=False,
            header=False,
            startrow=6  # data starts at row 7
        )

        workbook = writer.book
        sheet = writer.sheets["Document Register"]

        # Add specified text into specified cells
        sheet["A1"] = "Proceeding Title:"
        sheet["A2"] = "Court Proceeding No:"
        sheet["A3"] = "Party Creating Document:"
        sheet["F3"] = "Date of Document:"

        # 1) Black header row
        black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)

        for col_idx, header_name in enumerate(custom_headers, start=1):
            cell = sheet.cell(row=6, column=col_idx)
            cell.value = header_name
            cell.fill = black_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 2) Column widths (adjust if needed)
        column_widths = [15, 15, 15, 40, 60, 15, 15, 20, 20, 40, 40]
        for col_idx, width in enumerate(column_widths, start=1):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = width

        # 3) Wrap text in data area
        data_end_row = 6 + len(df)
        for row in range(7, data_end_row + 1):
            for col in range(1, len(custom_headers) + 1):
                cell = sheet.cell(row=row, column=col)
                cell.alignment = Alignment(wrap_text=True)

    print(f"✅ Document Register saved to: {output_filename}")